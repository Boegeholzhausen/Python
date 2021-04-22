import openpyxl
import os

os.chdir(r"C:\Python\automate_the_boring_stuff\Chapter 13 - Working with Excelsheets")

workbook = openpyxl.load_workbook("example.xlsx")

sheet = workbook.get_sheet_by_name("Sheet1")
print(sheet)

cell = sheet["A1"]

print(cell.value)
print(sheet["A2"].value)

print("\n")

for i in range (1, 8):
    print(sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value)