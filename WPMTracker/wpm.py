import os, re, pyperclip
import openpyxl
import time, datetime
from datetime import time
from datetime import date


# nach Values suchen
#text = '''
#94 WPM
#(Wörter pro Minute)
#Top 20 %
#Tastenanschläge	(468 | 28) 496
#Genauigkeit	93.41%
#Korrekte Wörter	86
#Falsche Wörter	5
#'''
text = pyperclip.paste()

wpm = re.compile(r'(\d|\d\d|\d\d\d) WPM')
allwpm = wpm.findall(text)

hits = re.compile(r'(\d|\d\d|\d\d\d) \|')
allhits = hits.findall(text)

failed = re.compile(r'(\d|\d\d|\d\d\d)\)')
allfailed = failed.findall(text)

totalhits = int(allhits[0]) + int(allfailed[0])

acc = re.compile(r'(\d\d.\d\d)\%')
allacc = acc.findall(text)

right = re.compile(r'Korrekte Wörter	(.*)')
allright = right.findall(text)

wrong = re.compile(r'Falsche Wörter	(.*)')
allwrong = wrong.findall(text)

# Values ausgeben (optional)
print(f"""
WPM: {allwpm[0]}
Tastenanschläge: {totalhits} ({allhits[0]}|{allfailed[0]})
Genauigkeit: {allacc[0]}%
Korrekte Wörter: {allright[0]}
Falsche Wörter: {allwrong[0]}
""")


# Zeit
now = datetime.datetime.now()
print(now.strftime("%d.%m.%Y %H:%M:%S"))


# Excelfile öffnen
os.chdir(r"C:\Python\WPMTracker")
workbook = openpyxl.load_workbook("WPM.xlsx")
sheet = workbook["WPM"]


# Überprüfen, wo letzte gefüllte Zeile ist
start = 2
for i in range(start, 100000000):
    if sheet.cell(row=i, column=1).value == None:
        sheet.cell(row=i, column=1).value = i-1
        sheet.cell(row=i, column=2).value = now
        sheet.cell(row=i, column=3).value = int(allwpm[0])
        sheet.cell(row=i, column=4).value = totalhits
        sheet.cell(row=i, column=5).value = int(allhits[0])
        sheet.cell(row=i, column=6).value = int(allfailed[0])
        sheet.cell(row=i, column=7).value = allacc[0]
        sheet.cell(row=i, column=8).value = int(allright[0])
        sheet.cell(row=i, column=9).value = int(allwrong[0])
        break
    else:
        start += 1

workbook.save("WPM.xlsx")





