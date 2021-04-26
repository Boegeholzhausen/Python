import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import os


# open sheet 
os.chdir(r"C:\Python\WPMTracker")
workbook = openpyxl.load_workbook("WPM.xlsx")
sheet = workbook["WPM"]

# Data import from Excel
nrvalues = []
start = 2
for i in range(start, 1000):
    if sheet.cell(row=i, column=1).value != None:
        nrvalues.append(sheet.cell(row=i, column=1).value)
        start += 1

wpmvalues = []
start2 = 2
for i in range(start2, 1000):
    if sheet.cell(row=i, column=3).value != None:
        wpmvalues.append(sheet.cell(row=i, column=3).value)
        start += 1

#accvalues = []
#start2 = 2
#for i in range(start2, 1000):
#    if sheet.cell(row=i, column=4).value != None:
#        accvalues.append(sheet.cell(row=i, column=4).value)
#        start += 1
#print(accvalues)


# Data
dfwpm=pd.DataFrame({'xvalues': nrvalues, 'yvalues': wpmvalues})
#dfacc=pd.DataFrame({'xaccvalues': nrvalues, 'yaccvalues': accvalues})

# Plot
plt.style.use("seaborn")
plt.figure(figsize=(22,6))
plt.rcParams["figure.figsize"]
plt.plot('xvalues', 'yvalues', color="#2F78BB", linewidth=3, label='WPM Tracker (WPM)', data=dfwpm)
plt.title("WPM Tracker")
plt.xlabel("Nr.")
plt.ylabel("WPM")
plt.grid(True)
plt.tight_layout()


#plt.style.use("seaborn")
#plt.figure(figsize=(20,5))
#plt.rcParams["figure.figsize"]
#plt.plot('xaccvalues', 'yaccvalues', color="#2F78BB", linewidth=3, label='WPM Tracker (Accuracy)', data=dfacc)
#plt.title("WPM Tracker")
#plt.xlabel("Nr.")
#plt.ylabel("WPM")
#plt.grid(True)
#plt.tight_layout()
#


plt.show()
