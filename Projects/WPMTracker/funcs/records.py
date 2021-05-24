import openpyxl
from tkinter import *
from tkinter.ttk import *
import os

os.chdir(r"C:\Python\Projects\WPMTracker")


def wpm():
    # open sheet easy
    workbook = openpyxl.load_workbook("WPM.xlsx")
    sheet = workbook["WPM"]

    # Data import from Excel easy
    global maxwpm
    wpmvalues = []

    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=3).value != None:
            wpmvalues.append(sheet.cell(row=i, column=3).value)
            start += 1
    maxwpm = max(wpmvalues)


def hwpm():
    # open sheet easy
    workbook = openpyxl.load_workbook("hWPM.xlsx")
    sheet = workbook["WPM"]

    # Data import from Excel easy
    global maxhwpm
    hwpmvalues = []

    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=3).value != None:
            hwpmvalues.append(sheet.cell(row=i, column=3).value)
            start += 1
    maxhwpm = max(hwpmvalues)


def acc():
    # open sheet easy
    workbook = openpyxl.load_workbook("WPM.xlsx")
    sheet = workbook["WPM"]

    # Data import from Excel easy
    global maxacc
    accvalues = []

    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=4).value != None:
            accvalues.append(sheet.cell(row=i, column=4).value)
            start += 1
    maxacc = max(accvalues)


def hacc():
    # open sheet easy
    workbook = openpyxl.load_workbook("hWPM.xlsx")
    sheet = workbook["WPM"]

    # Data import from Excel easy
    global maxhacc
    haccvalues = []

    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=4).value != None:
            haccvalues.append(sheet.cell(row=i, column=4).value)
            start += 1
    maxhacc = max(haccvalues)


def hits():
    # open sheet easy
    workbook = openpyxl.load_workbook("WPM.xlsx")
    sheet = workbook["WPM"]

    # Data import from Excel easy
    global maxhits
    hitsvalues = []

    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=5).value != None:
            hitsvalues.append(sheet.cell(row=i, column=5).value)
            start += 1
    maxhits = max(hitsvalues)


def hhits():
    # open sheet easy
    workbook = openpyxl.load_workbook("hWPM.xlsx")
    sheet = workbook["WPM"]

    # Data import from Excel easy
    global maxhhits
    hhitsvalues = []

    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=5).value != None:
            hhitsvalues.append(sheet.cell(row=i, column=5).value)
            start += 1
    maxhhits = max(hhitsvalues)