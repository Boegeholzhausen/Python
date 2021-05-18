import os, openpyxl, re, pyperclip
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, colors
from openpyxl.cell import Cell
import datetime


# Farben
red = PatternFill(start_color='DD5151', end_color='DD5151', fill_type='solid')
green = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
blue = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
grey = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
lightgrey = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')                
rightbot = Border(right=Side(style='thick'), bottom=Side(style='thin'))
fontb = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')


# Values in Excel eintragen [easy]
def fill_excel_easy():
    now = datetime.datetime.now()
    # Path wählen und Excelfile öffnen
    os.chdir(r"C:\Python\Projects\WPMTracker")
    workbook = openpyxl.load_workbook("WPM.xlsx")
    sheet = workbook["WPM"]
    
    text = pyperclip.paste()

    wpm = re.compile(r'(\d|\d\d|\d\d\d) WPM')
    allwpm = wpm.findall(text)

    hits = re.compile(r'(\d|\d\d|\d\d\d) \|')
    allhits = hits.findall(text)

    failed = re.compile(r'(\d|\d\d|\d\d\d)\)')
    allfailed = failed.findall(text)

    totalhits = int(allhits[0]) + int(allfailed[0])

    acc = re.compile(r'(\d\d.\d|\d\d.\d\d)\%')
    allacc = acc.findall(text)

    right = re.compile(r'Korrekte Wörter	(.*)')
    allright = right.findall(text)

    wrong = re.compile(r'Falsche Wörter	(.*)')
    allwrong = wrong.findall(text)


    # Überprüfen, wo letzte gefüllte Zeile ist, die dann mit Werten füllen
    start = 2
    for i in range(start, 100000000):
        if sheet.cell(row=i, column=1).value == None:
            sheet.cell(row=i, column=1).alignment = Alignment(horizontal='left')
            sheet.cell(row=i, column=1).font = font
            sheet.cell(row=i, column=1).fill = white
            sheet.cell(row=i, column=1).border = rightbot
            sheet.cell(row=i, column=1).value = i-1

            sheet.cell(row=i, column=2).font = font
            sheet.cell(row=i, column=2).fill = lightgrey
            sheet.cell(row=i, column=2).border = rightbot
            sheet.cell(row=i, column=2).value = now

            sheet.cell(row=i, column=3).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=3).font = fontb
            sheet.cell(row=i, column=3).fill = blue
            sheet.cell(row=i, column=3).border = rightbot
            sheet.cell(row=i, column=3).value = int(allwpm[0])

            # sheet.cell(row=i, column=4).style = "Percent"
            sheet.cell(row=i, column=4).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=4).font = font
            sheet.cell(row=i, column=4).fill = lightgrey
            sheet.cell(row=i, column=4).border = rightbot
            sheet.cell(row=i, column=4).value = allacc[0] 

            sheet.cell(row=i, column=5).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=5).font = font
            sheet.cell(row=i, column=5).fill = grey
            sheet.cell(row=i, column=5).border = rightbot
            sheet.cell(row=i, column=5).value = totalhits

            sheet.cell(row=i, column=6).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=6).font = font
            sheet.cell(row=i, column=6).fill = green
            sheet.cell(row=i, column=6).border = rightbot
            sheet.cell(row=i, column=6).value = int(allhits[0])

            sheet.cell(row=i, column=7).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=7).font = font
            sheet.cell(row=i, column=7).fill = red
            sheet.cell(row=i, column=7).border = rightbot
            sheet.cell(row=i, column=7).value = int(allfailed[0])

            sheet.cell(row=i, column=8).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=8).font = font
            sheet.cell(row=i, column=8).fill = green
            sheet.cell(row=i, column=8).border = rightbot
            sheet.cell(row=i, column=8).value = int(allright[0])

            sheet.cell(row=i, column=9).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=9).font = font
            sheet.cell(row=i, column=9).fill = red
            sheet.cell(row=i, column=9).border = rightbot
            sheet.cell(row=i, column=9).value = int(allwrong[0])
            break
        else:
            start += 1
    workbook.save("WPM.xlsx")

# Values in Excel eintragen [hard]
def fill_excel_hard():
    now = datetime.datetime.now()
    # Path wählen und Excelfile öffnen
    os.chdir(r"C:\Python\Projects\WPMTracker")
    workbook = openpyxl.load_workbook("hWPM.xlsx")
    sheet = workbook["WPM"]
    
    text = pyperclip.paste()

    wpm = re.compile(r'(\d|\d\d|\d\d\d) WPM')
    allwpm = wpm.findall(text)

    hits = re.compile(r'(\d|\d\d|\d\d\d) \|')
    allhits = hits.findall(text)

    failed = re.compile(r'(\d|\d\d|\d\d\d)\)')
    allfailed = failed.findall(text)

    totalhits = int(allhits[0]) + int(allfailed[0])

    acc = re.compile(r'(\d\d.\d|\d\d.\d\d)\%')
    allacc = acc.findall(text)

    right = re.compile(r'Korrekte Wörter	(.*)')
    allright = right.findall(text)

    wrong = re.compile(r'Falsche Wörter	(.*)')
    allwrong = wrong.findall(text)


    # Überprüfen, wo letzte gefüllte Zeile ist, die dann mit Werten füllen
    start = 2
    for i in range(start, 100000000):
        if sheet.cell(row=i, column=1).value == None:
            sheet.cell(row=i, column=1).alignment = Alignment(horizontal='left')
            sheet.cell(row=i, column=1).font = font
            sheet.cell(row=i, column=1).fill = white
            sheet.cell(row=i, column=1).border = rightbot
            sheet.cell(row=i, column=1).value = i-1

            sheet.cell(row=i, column=2).font = font
            sheet.cell(row=i, column=2).fill = lightgrey
            sheet.cell(row=i, column=2).border = rightbot
            sheet.cell(row=i, column=2).value = now

            sheet.cell(row=i, column=3).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=3).font = fontb
            sheet.cell(row=i, column=3).fill = blue
            sheet.cell(row=i, column=3).border = rightbot
            sheet.cell(row=i, column=3).value = int(allwpm[0])

            # sheet.cell(row=i, column=4).style = "Percent"
            sheet.cell(row=i, column=4).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=4).font = font
            sheet.cell(row=i, column=4).fill = lightgrey
            sheet.cell(row=i, column=4).border = rightbot
            sheet.cell(row=i, column=4).value = allacc[0] 

            sheet.cell(row=i, column=5).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=5).font = font
            sheet.cell(row=i, column=5).fill = grey
            sheet.cell(row=i, column=5).border = rightbot
            sheet.cell(row=i, column=5).value = totalhits

            sheet.cell(row=i, column=6).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=6).font = font
            sheet.cell(row=i, column=6).fill = green
            sheet.cell(row=i, column=6).border = rightbot
            sheet.cell(row=i, column=6).value = int(allhits[0])

            sheet.cell(row=i, column=7).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=7).font = font
            sheet.cell(row=i, column=7).fill = red
            sheet.cell(row=i, column=7).border = rightbot
            sheet.cell(row=i, column=7).value = int(allfailed[0])

            sheet.cell(row=i, column=8).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=8).font = font
            sheet.cell(row=i, column=8).fill = green
            sheet.cell(row=i, column=8).border = rightbot
            sheet.cell(row=i, column=8).value = int(allright[0])

            sheet.cell(row=i, column=9).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=9).font = font
            sheet.cell(row=i, column=9).fill = red
            sheet.cell(row=i, column=9).border = rightbot
            sheet.cell(row=i, column=9).value = int(allwrong[0])
            break
        else:
            start += 1
    workbook.save("hWPM.xlsx")