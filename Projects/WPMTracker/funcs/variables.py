# Farben und Muster für Excel
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, colors
from openpyxl.styles.borders import Border, Side
import datetime

# Farben für Excel
red = PatternFill(start_color='DD5151', end_color='DD5151', fill_type='solid') # [hard buttons]
green = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid') # [easy buttons]
blue = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid') # [background]
grey = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
lightgrey = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')                
rightbot = Border(right=Side(style='thick'), bottom=Side(style='thin'))
fontb = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')