import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import os

os.chdir(r"C:\Python\Projects\WPMTracker")


def graph():
    global dfwpm
    global dfacc
    global dfhits
    # open sheet easy
    workbook = openpyxl.load_workbook("WPM.xlsx")
    sheet = workbook["WPM"]

    # Data import from Excel easy
    nrvalues = []
    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=1).value != None:
            nrvalues.append(sheet.cell(row=i, column=1).value)
            start += 1

    wpmvalues = []
    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=3).value != None:
            wpmvalues.append(sheet.cell(row=i, column=3).value)
            start += 1

    accvalues = []
    newacc = []
    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=4).value != None:
            accvalues.append(sheet.cell(row=i, column=4).value)
            start += 1
            word = accvalues[i-2]
            newacc.append(float(word))

    hitsvalues = []
    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=5).value != None:
            hitsvalues.append(sheet.cell(row=i, column=5).value)
            start += 1

    # Data
    dfwpm=pd.DataFrame({'xvalues': nrvalues, 'yvalues': wpmvalues})
    dfacc=pd.DataFrame({'xvalues': nrvalues, 'yvalues': newacc})
    dfhits=pd.DataFrame({'xvalues': nrvalues, 'yvalues': hitsvalues})
    return dfwpm, dfacc, dfhits


def hgraph():
    global dfhwpm
    global dfhacc
    global dfhhits
    # open sheet easy
    workbook = openpyxl.load_workbook("hWPM.xlsx")
    sheet = workbook["WPM"]

    # Data import from Excel easy
    nrvalues = []
    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=1).value != None:
            nrvalues.append(sheet.cell(row=i, column=1).value)
            start += 1

    wpmvalues = []
    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=3).value != None:
            wpmvalues.append(sheet.cell(row=i, column=3).value)
            start += 1

    accvalues = []
    newacc = []
    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=4).value != None:
            accvalues.append(sheet.cell(row=i, column=4).value)
            start += 1
            word = accvalues[i-2]
            newacc.append(float(word))

    hitsvalues = []
    start = 2
    for i in range(start, 100000):
        if sheet.cell(row=i, column=5).value != None:
            hitsvalues.append(sheet.cell(row=i, column=5).value)
            start += 1

    # Data
    dfhwpm=pd.DataFrame({'xvalues': nrvalues, 'yvalues': wpmvalues})
    dfhacc=pd.DataFrame({'xvalues': nrvalues, 'yvalues': newacc})
    dfhhits=pd.DataFrame({'xvalues': nrvalues, 'yvalues': hitsvalues})
    return dfhwpm, dfhacc, dfhhits




#     # Plot WPM
#     plt.style.use("seaborn")
#     plt.figure(figsize=(23,5))
#     plt.plot('xvalues', 'yvalues', color="#2F78BB", linewidth=4, label='WPM Tracker (WPM)', data=dfwpm)
#     plt.title("WPM Tracker (WPM)")
#     plt.xlabel("Nr.")
#     plt.ylabel("WPM")
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()
# 
#     # Plot Acc
#     plt.style.use("seaborn")
#     plt.figure(figsize=(23,5))
#     plt.plot('xvalues', 'yvalues', color="#2F78BB", linewidth=3, label='WPM Tracker (Accuracy)', data=dfacc)
#     plt.title("WPM Tracker (Accuracy)")
#     plt.xlabel("Nr.")
#     plt.ylabel("%")
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()
# 
#     # Plot Tastenanschläge
#     plt.style.use("seaborn")
#     plt.figure(figsize=(23,5))
#     plt.plot('xvalues', 'yvalues', color="#2F78BB", linewidth=3, label='WPM Tracker (Tastenanschläge)', data=dfhits)
#     plt.title("WPM Tracker (Tastenanschläge)")
#     plt.xlabel("Nr.")
#     plt.ylabel("Tastenanschläge")
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()
    



# ## Graphs
# def wpm():
#     # open sheet easy
#     workbook = openpyxl.load_workbook("WPM.xlsx")
#     sheet = workbook["WPM"]
# 
#     # Data import from Excel easy
#     nrvalues = []
#     start = 2
#     for i in range(start, 100000):
#         if sheet.cell(row=i, column=1).value != None:
#             nrvalues.append(sheet.cell(row=i, column=1).value)
#             start += 1
# 
#     wpmvalues = []
#     start = 2
#     for i in range(start, 100000):
#         if sheet.cell(row=i, column=3).value != None:
#             wpmvalues.append(sheet.cell(row=i, column=3).value)
#             start += 1
# 
#     # Data
#     dfwpm=pd.DataFrame({'xvalues': nrvalues, 'yvalues': wpmvalues})
# 
#     # Plot WPM
#     plt.style.use("seaborn")
#     plt.figure(figsize=(23,5))
#     plt.plot('xvalues', 'yvalues', color="#2F78BB", linewidth=4, label='WPM Tracker (WPM)', data=dfwpm)
#     plt.title("WPM Tracker (WPM)")
#     plt.xlabel("Nr.")
#     plt.ylabel("WPM")
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()
# 
# 
# def acc():
#     # open sheet easy
#     workbook = openpyxl.load_workbook("WPM.xlsx")
#     sheet = workbook["WPM"]
# 
#     # Data import from Excel easy
#     nrvalues = []
#     start = 2
#     for i in range(start, 100000):
#         if sheet.cell(row=i, column=1).value != None:
#             nrvalues.append(sheet.cell(row=i, column=1).value)
#             start += 1
# 
# 
#     accvalues = []
#     newacc = []
#     start = 2
#     for i in range(start, 100000):
#         if sheet.cell(row=i, column=4).value != None:
#             accvalues.append(sheet.cell(row=i, column=4).value)
#             start += 1
#             word = accvalues[i-2]
#             newacc.append(float(word))
# 
#     # Data
#     dfacc=pd.DataFrame({'xvalues': nrvalues, 'yvalues': newacc})
# 
#     # Plot Acc
#     plt.style.use("seaborn")
#     plt.figure(figsize=(23,5))
#     plt.plot('xvalues', 'yvalues', color="#2F78BB", linewidth=3, label='WPM Tracker (Accuracy)', data=dfacc)
#     plt.title("WPM Tracker (Accuracy)")
#     plt.xlabel("Nr.")
#     plt.ylabel("%")
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()
# 
# 
# def hits():
#     # open sheet easy
#     workbook = openpyxl.load_workbook("WPM.xlsx")
#     sheet = workbook["WPM"]
# 
#     # Data import from Excel easy
#     nrvalues = []
#     start = 2
#     for i in range(start, 100000):
#         if sheet.cell(row=i, column=1).value != None:
#             nrvalues.append(sheet.cell(row=i, column=1).value)
#             start += 1
# 
#     hitsvalues = []
#     start = 2
#     for i in range(start, 100000):
#         if sheet.cell(row=i, column=5).value != None:
#             hitsvalues.append(sheet.cell(row=i, column=5).value)
#             start += 1
# 
#     # Data
#     dfhits=pd.DataFrame({'xvalues': nrvalues, 'yvalues': hitsvalues})
# 
#     # Plot Tastenanschläge
#     plt.style.use("seaborn")
#     plt.figure(figsize=(23,5))
#     plt.plot('xvalues', 'yvalues', color="#2F78BB", linewidth=3, label='WPM Tracker (Tastenanschläge)', data=dfhits)
#     plt.title("WPM Tracker (Tastenanschläge)")
#     plt.xlabel("Nr.")
#     plt.ylabel("Tastenanschläge")
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()
# 
# 
# def hwpm():
#     # open sheet hard
#     hworkbook = openpyxl.load_workbook("hWPM.xlsx")
#     hsheet = hworkbook["WPM"]
# 
#     hnrvalues = []
#     start = 2
#     for i in range(start, 100000):
#         if hsheet.cell(row=i, column=1).value != None:
#             hnrvalues.append(hsheet.cell(row=i, column=1).value)
#             start += 1
# 
#     # Data import from Excel hard
#     hwpmvalues = []
#     start = 2
#     for i in range(start, 100000):
#         if hsheet.cell(row=i, column=3).value != None:
#             hwpmvalues.append(hsheet.cell(row=i, column=3).value)
#             start += 1
# 
#     # Data
#     dfhwpm=pd.DataFrame({'xvalues': hnrvalues, 'yvalues': hwpmvalues})
#     
# 
#     # Plot WPM [hard]
#     plt.style.use("seaborn")
#     plt.figure(figsize=(23,5))
#     plt.plot('xvalues', 'yvalues', color="#2F78BB", linewidth=3, label='WPM Tracker (WPM) [Hard]', data=dfhwpm)
#     plt.title("WPM Tracker (WPM) [Hard]")
#     plt.xlabel("Nr.")
#     plt.ylabel("WPM")
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()
# 
# 
# def hacc():
#     # open sheet hard
#     hworkbook = openpyxl.load_workbook("hWPM.xlsx")
#     hsheet = hworkbook["WPM"]
# 
#     hnrvalues = []
#     start = 2
#     for i in range(start, 100000):
#         if hsheet.cell(row=i, column=1).value != None:
#             hnrvalues.append(hsheet.cell(row=i, column=1).value)
#             start += 1
# 
#     # Data import from Excel hard
#     haccvalues = []
#     hnewacc = []
#     start = 2
#     for i in range(start, 100000):
#         if hsheet.cell(row=i, column=4).value != None:
#             haccvalues.append(hsheet.cell(row=i, column=4).value)
#             start += 1
#             word = haccvalues[i-2]
#             hnewacc.append(float(word))
# 
#     # Data
#     dfhacc=pd.DataFrame({'xvalues': hnrvalues, 'yvalues': hnewacc})
# 
#     # Plot Acc [hard]
#     plt.style.use("seaborn")
#     plt.figure(figsize=(23,5))
#     plt.plot('xvalues', 'yvalues', color="#2F78BB", linewidth=3, label='WPM Tracker (Accuracy) [Hard]', data=dfhacc)
#     plt.title("WPM Tracker (Accuracy) [Hard]")
#     plt.xlabel("Nr.")
#     plt.ylabel("%")
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()
# 
# 
# def hhits():
#     # open sheet hard
#     hworkbook = openpyxl.load_workbook("hWPM.xlsx")
#     hsheet = hworkbook["WPM"]
# 
#     hnrvalues = []
#     start = 2
#     for i in range(start, 100000):
#         if hsheet.cell(row=i, column=1).value != None:
#             hnrvalues.append(hsheet.cell(row=i, column=1).value)
#             start += 1
# 
#     # Data import from Excel hard
#     hhitsvalues = []
#     start = 2
#     for i in range(start, 100000):
#         if hsheet.cell(row=i, column=5).value != None:
#             hhitsvalues.append(hsheet.cell(row=i, column=5).value)
#             start += 1
# 
#     # Data
#     dfhhits=pd.DataFrame({'xvalues': hnrvalues, 'yvalues': hhitsvalues})
#     
# 
#     # Plot Tastenanschläge [hard]
#     plt.style.use("seaborn")
#     plt.figure(figsize=(23,5))
#     plt.plot('xvalues', 'yvalues', color="#2F78BB", linewidth=3, label='WPM Tracker (Tastenanschläge) [Hard]', data=dfhhits)
#     plt.title("WPM Tracker (Tastenanschläge) [Hard]")
#     plt.xlabel("Nr.")
#     plt.ylabel("Tastenanschläge")
#     plt.grid(True)
#     plt.tight_layout()
#     plt.show()
# 