import os, re, pyperclip
import openpyxl
import time, datetime
from datetime import time
from datetime import date
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment, colors
from openpyxl.cell import Cell
import variables as var
import copy_value as cv


# Values Kopieren, Werte finden und in Excel passend speichern [easy]
def getValues():
    # Kopiere Values von der Seite
    cv.copy()

    # Path wählen und Excelfile öffnen
    os.chdir(r"C:\Python\Projects\WPMTracker")
    workbook = openpyxl.load_workbook("WPM.xlsx")
    sheet = workbook["WPM"]
    text = pyperclip.paste()

    wpm = re.compile(r'(\d|\d\d|\d\d\d) WPM')
    allwpm = wpm.findall(text)

    hits = re.compile(r'(\d|\d\d|\d\d\d) \|')
    allhits = hits.findall(text)
    
    failed = re.compile(r'(\d|\d\d|\d\d\d)\)')
    allfailed = failed.findall(text)
    
    totalhits = int(allhits[0]) + int(allfailed[0])
    
    acc = re.compile(r'(\d\d.\d|\d\d.\d\d)\%')
    allacc = acc.findall(text)
    
    right = re.compile(r'Korrekte Wörter	(.*)')
    allright = right.findall(text)
    
    wrong = re.compile(r'Falsche Wörter	(.*)')
    allwrong = wrong.findall(text)
    

    # Überprüfen, wo letzte gefüllte Zeile ist, die dann mit Werten füllen
    start = 2
    for i in range(start, 100000000):
        if sheet.cell(row=i, column=1).value == None:
            sheet.cell(row=i, column=1).alignment = Alignment(horizontal='left')
            sheet.cell(row=i, column=1).font = var.font
            sheet.cell(row=i, column=1).fill = var.white
            sheet.cell(row=i, column=1).border = var.rightbot
            sheet.cell(row=i, column=1).value = i-1
    
            sheet.cell(row=i, column=2).font = var.font
            sheet.cell(row=i, column=2).fill = var.lightgrey
            sheet.cell(row=i, column=2).border = var.rightbot
            sheet.cell(row=i, column=2).value = var.now
    
            sheet.cell(row=i, column=3).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=3).font = var.fontb
            sheet.cell(row=i, column=3).fill = var.blue
            sheet.cell(row=i, column=3).border = var.rightbot
            sheet.cell(row=i, column=3).value = int(allwpm[0])
    
            # sheet.cell(row=i, column=4).style = "Percent"
            sheet.cell(row=i, column=4).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=4).font = var.font
            sheet.cell(row=i, column=4).fill = var.lightgrey
            sheet.cell(row=i, column=4).border = var.rightbot
            sheet.cell(row=i, column=4).value = allacc[0] 
    
            sheet.cell(row=i, column=5).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=5).font = var.font
            sheet.cell(row=i, column=5).fill = var.grey
            sheet.cell(row=i, column=5).border = var.rightbot
            sheet.cell(row=i, column=5).value = totalhits
    
            sheet.cell(row=i, column=6).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=6).font = var.font
            sheet.cell(row=i, column=6).fill = var.green
            sheet.cell(row=i, column=6).border = var.rightbot
            sheet.cell(row=i, column=6).value = int(allhits[0])
    
            sheet.cell(row=i, column=7).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=7).font = var.font
            sheet.cell(row=i, column=7).fill = var.red
            sheet.cell(row=i, column=7).border = var.rightbot
            sheet.cell(row=i, column=7).value = int(allfailed[0])
    
            sheet.cell(row=i, column=8).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=8).font = var.font
            sheet.cell(row=i, column=8).fill = var.green
            sheet.cell(row=i, column=8).border = var.rightbot
            sheet.cell(row=i, column=8).value = int(allright[0])
    
            sheet.cell(row=i, column=9).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=9).font = var.font
            sheet.cell(row=i, column=9).fill = var.red
            sheet.cell(row=i, column=9).border = var.rightbot
            sheet.cell(row=i, column=9).value = int(allwrong[0])
            break
        else:
            start += 1
    
    workbook.save("WPM.xlsx")


# Values Kopieren, Werte finden und in Excel passend speichern [hard]
def gethValues():
    # Kopiere Values von der Seite
    cv.copy()

    # Path wählen und Excelfile öffnen
    os.chdir(r"C:\Python\Projects\WPMTracker")
    workbook = openpyxl.load_workbook("hWPM.xlsx")
    sheet = workbook["WPM"]
    
    text = pyperclip.paste()

    wpm = re.compile(r'(\d|\d\d|\d\d\d) WPM')
    allwpm = wpm.findall(text)

    hits = re.compile(r'(\d|\d\d|\d\d\d) \|')
    allhits = hits.findall(text)

    failed = re.compile(r'(\d|\d\d|\d\d\d)\)')
    allfailed = failed.findall(text)

    totalhits = int(allhits[0]) + int(allfailed[0])

    acc = re.compile(r'(\d\d.\d|\d\d.\d\d)\%')
    allacc = acc.findall(text)

    right = re.compile(r'Korrekte Wörter	(.*)')
    allright = right.findall(text)

    wrong = re.compile(r'Falsche Wörter	(.*)')
    allwrong = wrong.findall(text)


    # Überprüfen, wo letzte gefüllte Zeile ist, die dann mit Werten füllen
    start = 2
    for i in range(start, 100000000):
        if sheet.cell(row=i, column=1).value == None:
            sheet.cell(row=i, column=1).alignment = Alignment(horizontal='left')
            sheet.cell(row=i, column=1).font = var.font
            sheet.cell(row=i, column=1).fill = var.white
            sheet.cell(row=i, column=1).border = var.rightbot
            sheet.cell(row=i, column=1).value = i-1

            sheet.cell(row=i, column=2).font = var.font
            sheet.cell(row=i, column=2).fill = var.lightgrey
            sheet.cell(row=i, column=2).border = var.rightbot
            sheet.cell(row=i, column=2).value = var.now

            sheet.cell(row=i, column=3).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=3).font = var.fontb
            sheet.cell(row=i, column=3).fill = var.blue
            sheet.cell(row=i, column=3).border = var.rightbot
            sheet.cell(row=i, column=3).value = int(allwpm[0])

            # sheet.cell(row=i, column=4).style = "Percent"
            sheet.cell(row=i, column=4).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=4).font = var.font
            sheet.cell(row=i, column=4).fill = var.lightgrey
            sheet.cell(row=i, column=4).border = var.rightbot
            sheet.cell(row=i, column=4).value = allacc[0] 

            sheet.cell(row=i, column=5).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=5).font = var.font
            sheet.cell(row=i, column=5).fill = var.grey
            sheet.cell(row=i, column=5).border = var.rightbot
            sheet.cell(row=i, column=5).value = totalhits

            sheet.cell(row=i, column=6).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=6).font = var.font
            sheet.cell(row=i, column=6).fill = var.green
            sheet.cell(row=i, column=6).border = var.rightbot
            sheet.cell(row=i, column=6).value = int(allhits[0])

            sheet.cell(row=i, column=7).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=7).font = var.font
            sheet.cell(row=i, column=7).fill = var.red
            sheet.cell(row=i, column=7).border = var.rightbot
            sheet.cell(row=i, column=7).value = int(allfailed[0])

            sheet.cell(row=i, column=8).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=8).font = var.font
            sheet.cell(row=i, column=8).fill = var.green
            sheet.cell(row=i, column=8).border = var.rightbot
            sheet.cell(row=i, column=8).value = int(allright[0])

            sheet.cell(row=i, column=9).alignment = Alignment(horizontal='center')
            sheet.cell(row=i, column=9).font = var.font
            sheet.cell(row=i, column=9).fill = var.red
            sheet.cell(row=i, column=9).border = var.rightbot
            sheet.cell(row=i, column=9).value = int(allwrong[0])
            break
        else:
            start += 1

    workbook.save("hWPM.xlsx")

if __name__ == "__main__":
    getValues()
    gethValues()