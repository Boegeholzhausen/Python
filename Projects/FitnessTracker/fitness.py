import os, datetime
import tkinter as tk
from PIL import Image, ImageTk
from tkinter.filedialog import Toplevel
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import funcs.get_data as gd
import openpyxl


# chance directory
os.chdir(r"C:\Python\Projects\FitnessTracker")

workbook = openpyxl.load_workbook("Workout.xlsx")
sheet1 = workbook["Übung_1"]
sheet2 = workbook["Übung_2"]
sheet3 = workbook["Übung_3"]
sheet4 = workbook["Übung_4"]        
sheet5 = workbook["Übung_5"]
sheet6 = workbook["Übung_6"]
sheet7 = workbook["Übung_7"]

root = tk.Tk()
root.title("Fitness Tracker")
root.iconbitmap(r"C:\Python\Projects\FitnessTracker\icons\fitness_icon.ico")
background = "#488953"
root.resizable(False, False)
root.geometry("+1590+150")

canvas = tk.Canvas(root, width=500, height=530, bg=background)
canvas.grid(columnspan=2, rowspan=2)

## Start Screen (chose workout)
# Logo
logo = Image.open(r"C:\Python\Projects\FitnessTracker\icons\fitness.png")
logo = ImageTk.PhotoImage(logo)
logo_label = tk.Label(root, image=logo, bg=background)
logo_label.image = logo
logo_label.grid(columnspan=2, column=0, row=0)

# Button für Trainingsplanauswahl
btnh = 2
btnw = 17
tblh = 1
tblw = 7
setw = 5
btnfontsize = 14
txtfontsize = 16
btncolor = "#3C58A0"


# Intro Buttons
workout1_btn = tk.Button(root, text="Ring Workout", command=lambda:[workout_1_tabel(), workout_1_graph()])
workout1_btn.grid(column=0, row=1)
workout1_btn.config(bg=btncolor, fg="white", font=("Calibri bold", btnfontsize), height=btnh, width=btnw, borderwidth=5)

workout2_btn = tk.Button(root, text="Muscle Growth", command=lambda:workout_2())
workout2_btn.grid(column=1, row=1)
workout2_btn.config(bg=btncolor, fg="white", font=("Calibri bold", btnfontsize), height=btnh, width=btnw, borderwidth=5)


## Tabelle aufrufen
def workout_1_tabel():
    global counter_x
    global counter_y
    workout_1_tabel = Toplevel(root)
    workout_1_tabel.title("Ring Workout")
    workout_1_tabel.iconbitmap(r"C:\Python\Projects\FitnessTracker\icons\fitness_icon.ico")
    workout_1_tabel.resizable(False, False)
    canvas_1 = tk.Canvas(workout_1_tabel, width=1090, height=530, bg=background)
    canvas_1.grid(columnspan=5, rowspan=8)
    workout_1_tabel.geometry("+490+150")

    ## Timer
    #def start_timer():
    #    time = -1
    #    run = False
    #    timer_lbl = tk.Label(workout_1_tabel, text="timer")
    #    timer_lbl.grid(column=4, row=4, rowspan=2, pady=10, padx=10)
    #    timer_lbl.config(bg=background, font=("Calibri bold", 14), fg="white", width=14)    

    ## Tabelle
    tablecolor = "white"
    textcolor = "black"
    # Obere Reihe
    plan = tk.Label(workout_1_tabel, text="Übung", borderwidth=2, relief="solid")
    plan.grid(columnspan=1, rowspan=1, column=0, row=0, sticky="S", padx=30, pady=20, ipady=15, ipadx=50)
    plan.config(bg=tablecolor, font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=tblw)

    set1 = tk.Label(workout_1_tabel, text="Set 1", borderwidth=2, relief="solid")
    set1.grid(columnspan=1, rowspan=1, column=1, row=0, sticky="S", pady=20, ipady=15, ipadx=50)
    set1.config(bg=tablecolor, font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)

    set2 = tk.Label(workout_1_tabel, text="Set 2", borderwidth=2, relief="solid")
    set2.grid(columnspan=1, rowspan=1, column=2, row=0, sticky="S", pady=20, ipady=15, ipadx=50)
    set2.config(bg=tablecolor, font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)

    set3 = tk.Label(workout_1_tabel, text="Set 3", borderwidth=2, relief="solid")
    set3.grid(columnspan=1, rowspan=1, column=3, row=0, sticky="S", pady=20, ipady=15, ipadx=50)
    set3.config(bg=tablecolor, font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)

    # 7 Übungen
    for i in range(1,8):
        exercise = tk.Label(workout_1_tabel, text=str(gd.exercise[i-1]), borderwidth=2, relief="solid")
        exercise.grid(columnspan=1, rowspan=1, column=0, row=i, sticky="N", pady=1, ipady=10, ipadx=50)
        exercise.config(bg=tablecolor, font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=tblw)

    # Reps Kästchen erstellen
    for i in range(1,8):
        for x in range(1,4):
            rep = tk.Label(workout_1_tabel, text=" ", relief="solid") # Hier die eingetragenen Reps ausgeben
            rep.grid(columnspan=1, rowspan=1, column=x, row=i, sticky="N", pady=1, ipady=10, ipadx=50)
            rep.config(bg=tablecolor, font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
            # Text, Farbe ändern wenn Start Button gedrückt

    def start_training():
        global entry_field
        start_btn = tk.Button(workout_1_tabel, text="Training")
        start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
        start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 

        # Eingabe für die Reps
        entry_btn = tk.Button(workout_1_tabel, text="Wiederholungen eingeben", command=lambda:[enter_value(counter_x, counter_y), make_list()])
        entry_btn.grid(rowspan=2, column=4, row=2, pady=10, padx=10)
        entry_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22)    

        entry_field = tk.Entry(workout_1_tabel) # Hier die Reps eintragen
        entry_field.grid(rowspan=2, column=4, row=3, padx=10)
        entry_field.config(bg="#424457", font=("Calibri bold", 14), fg="white", borderwidth=3, width=16)
        entry_field.bind("<Return>", (lambda event:[enter_value(counter_x, counter_y), make_list()]))


    workout = []
    def make_list():
        workout.append(entry_field.get())
        print(workout)

    # Values in Excel einfügen
    def upload():
        now = datetime.datetime.now()
        start_btn = tk.Button(workout_1_tabel, text="Uploaded to Excel") 
        start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
        start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22)

        os.chdir(r"C:\Python\Projects\FitnessTracker")
        workbook = openpyxl.load_workbook("Workout.xlsx")
        sheet1 = workbook["Übung_1"]
        sheet2 = workbook["Übung_2"]
        sheet3 = workbook["Übung_3"]
        sheet4 = workbook["Übung_4"]        
        sheet5 = workbook["Übung_5"]
        sheet6 = workbook["Übung_6"]
        sheet7 = workbook["Übung_7"]

        start = 2
        for i in range(start, 1000):
            if sheet1.cell(row=i, column=1).value == None:
                sheet1.cell(row=i, column=1).value = i-1
                sheet1.cell(row=i, column=2).value = now       
                sheet1.cell(row=i, column=3).value = int(workout[0])
                sheet1.cell(row=i, column=4).value = int(workout[1])
                sheet1.cell(row=i, column=5).value = int(workout[2])
                break
            else:
                start += 1

        start = 2
        for i in range(start, 1000):
            if sheet2.cell(row=i, column=1).value == None:    
                sheet2.cell(row=i, column=1).value = i-1    
                sheet2.cell(row=i, column=2).value = now          
                sheet2.cell(row=i, column=3).value = int(workout[3])
                sheet2.cell(row=i, column=4).value = int(workout[4])
                sheet2.cell(row=i, column=5).value = int(workout[5])
                break
            else:
                start += 1

        start = 2
        for i in range(start, 1000):
            if sheet3.cell(row=i, column=1).value == None:
                sheet3.cell(row=i, column=1).value = i-1
                sheet3.cell(row=i, column=2).value = now        
                sheet3.cell(row=i, column=3).value = int(workout[6])
                sheet3.cell(row=i, column=4).value = int(workout[7])
                sheet3.cell(row=i, column=5).value = int(workout[8])
                break
            else:
                start += 1

        start = 2
        for i in range(start, 1000):
            if sheet4.cell(row=i, column=1).value == None:
                sheet4.cell(row=i, column=1).value = i-1
                sheet4.cell(row=i, column=2).value = now                        
                sheet4.cell(row=i, column=3).value = int(workout[9])
                sheet4.cell(row=i, column=4).value = int(workout[10])
                sheet4.cell(row=i, column=5).value = int(workout[11])
                break
            else:
                start += 1

        start = 2
        for i in range(start, 1000):
            if sheet5.cell(row=i, column=1).value == None:
                sheet5.cell(row=i, column=1).value = i-1
                sheet5.cell(row=i, column=2).value = now        
                sheet5.cell(row=i, column=3).value = int(workout[12])
                sheet5.cell(row=i, column=4).value = int(workout[13])
                sheet5.cell(row=i, column=5).value = int(workout[14])
                break
            else:
                start += 1

        start = 2
        for i in range(start, 1000):
            if sheet6.cell(row=i, column=1).value == None:
                sheet6.cell(row=i, column=1).value = i-1
                sheet6.cell(row=i, column=2).value = now       
                sheet6.cell(row=i, column=3).value = int(workout[15])
                sheet6.cell(row=i, column=4).value = int(workout[16])
                sheet6.cell(row=i, column=5).value = int(workout[17])
                break
            else:
                start += 1

        start = 2
        for i in range(start, 1000):
            if sheet7.cell(row=i, column=1).value == None:
                sheet7.cell(row=i, column=1).value = i-1
                sheet7.cell(row=i, column=2).value = now      
                sheet7.cell(row=i, column=3).value = int(workout[18])
                sheet7.cell(row=i, column=4).value = int(workout[19])
                sheet7.cell(row=i, column=5).value = int(workout[20])
                break
            else:
                start += 1

        workbook.save("Workout.xlsx")

    # Eingabe der Values
    counter_x = 1
    counter_y = 1
    def enter_value(nr_col, nr_row):
        global counter_x
        global counter_y
        if nr_row == 1:
            if nr_col == 1:
                rep1 = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep1.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep1.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="20 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22)  
                counter_x += 1
            elif nr_col == 2:
                rep2 = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep2.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep2.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="19 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 3:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="18 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x = counter_x - 2
                counter_y += 1
        elif nr_row == 2:
            if nr_col == 1:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="17 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 2:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="16 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 3:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="15 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x = counter_x - 2
                counter_y += 1
        elif nr_row == 3:
            if nr_col == 1:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="14 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 2:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="13 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 3:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="12 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x = counter_x - 2
                counter_y += 1
        elif nr_row == 4:
            if nr_col == 1:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="11 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 2:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="10 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 3:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="9 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x = counter_x - 2
                counter_y += 1
        elif nr_row == 5:
            if nr_col == 1:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="8 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 2:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="7 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 3:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="6 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x = counter_x - 2
                counter_y += 1
        elif nr_row == 6:
            if nr_col == 1:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="5 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 2:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="4 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 3:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="3 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x = counter_x - 2
                counter_y += 1
        elif nr_row == 7:
            if nr_col == 1:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="2 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 2:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="1 Übungen noch")
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22) 
                counter_x += 1
            elif nr_col == 3:
                rep = tk.Label(workout_1_tabel, text=entry_field.get(), relief="solid") # Hier die eingetragenen Reps ausgeben
                rep.grid(columnspan=1, rowspan=1, column=nr_col, row=nr_row, sticky="N", pady=1, ipady=10, ipadx=50)
                rep.config(bg="#a0ff99", font=("Calibri bold", txtfontsize), fg=textcolor, height=tblh, width=setw)
                start_btn = tk.Button(workout_1_tabel, text="Finish & Upload", command=lambda: upload())
                start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
                start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22)
                entry_btn = tk.Button(workout_1_tabel, text="Wiederholungen eingeben")
                entry_btn.grid(rowspan=2, column=4, row=2, pady=10, padx=10)
                entry_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22)  
                

    ## Eingabe (rechts)
    # Ring Logo
    logo = Image.open(r"C:\Python\Projects\FitnessTracker\icons\gymrings.png")
    logo = ImageTk.PhotoImage(logo)
    logo_label = tk.Label(workout_1_tabel, image=logo, bg=background)
    logo_label.image = logo
    logo_label.grid(rowspan=3, column=4, row=0, pady=10)

    # Start Button
    start_btn = tk.Button(workout_1_tabel, text="Start Workout", command=lambda:start_training())
    start_btn.grid(column=4, row=5, rowspan=2, sticky="S", pady=10, padx=10)
    start_btn.config(bg=btncolor, font=("Calibri bold", 14), fg="white", borderwidth=3, width=22)  


# Graphen anzeigen
def workout_1_graph():
    workout_1_graph = Toplevel(root)
    workout_1_graph.title("Ring Workout")
    workout_1_graph.iconbitmap(r"C:\Python\Projects\FitnessTracker\icons\fitness_icon.ico")
    workout_1_graph.resizable(False, False)
    canvas_1 = tk.Canvas(workout_1_graph, width=1600, height=500, bg=background)
    canvas_1.grid(columnspan=2, rowspan=2)
    workout_1_graph.geometry("+490+718")

    # Graph anzeigen
    def change_Graph():
        cur_exer = variable.get()
        if cur_exer == "Pull-Up":
            gd.exer1()
            graph_exer = plt.figure(figsize=(15,4))
            graph_exer.patch.set_facecolor(background)
            line = FigureCanvasTkAgg(graph_exer, workout_1_graph)
            line.get_tk_widget().grid(columnspan=2, rowspan=1, column=0, row=1, padx=10, pady=10)
            plt.plot('xvalues', 'yvalues', color="#8380c2", linewidth=3, label=cur_exer + " [Set 1]", data = gd.dfset1)
            plt.plot('xvalues', 'yvalues', color="#33a3cc", linewidth=3, label=cur_exer + " [Set 2]", data = gd.dfset2)
            plt.plot('xvalues', 'yvalues', color="#09046c", linewidth=3, label=cur_exer + " [Set 3]", data = gd.dfset3)
            plt.xlabel("Nr.")
            plt.ylabel("Reps")
            plt.legend(loc='upper left')
            plt.grid(True)
            plt.tight_layout()
        elif cur_exer == "Dip":
            gd.exer2()
            graph_exer = plt.figure(figsize=(15,4))
            graph_exer.patch.set_facecolor(background)
            line = FigureCanvasTkAgg(graph_exer, workout_1_graph)
            line.get_tk_widget().grid(columnspan=2, rowspan=1, column=0, row=1, padx=10, pady=10)
            plt.plot('xvalues', 'yvalues', color="#8380c2", linewidth=3, label=cur_exer + " [Set 1]", data = gd.dfset1)
            plt.plot('xvalues', 'yvalues', color="#33a3cc", linewidth=3, label=cur_exer + " [Set 2]", data = gd.dfset2)
            plt.plot('xvalues', 'yvalues', color="#09046c", linewidth=3, label=cur_exer + " [Set 3]", data = gd.dfset3)
            plt.xlabel("Nr.")
            plt.ylabel("Reps")
            plt.legend(loc='upper left')
            plt.grid(True)
            plt.tight_layout()
        elif cur_exer == "Row":
            gd.exer3()
            graph_exer = plt.figure(figsize=(15,4))
            graph_exer.patch.set_facecolor(background)
            line = FigureCanvasTkAgg(graph_exer, workout_1_graph)
            line.get_tk_widget().grid(columnspan=2, rowspan=1, column=0, row=1, padx=10, pady=10)
            plt.plot('xvalues', 'yvalues', color="#8380c2", linewidth=3, label=cur_exer + " [Set 1]", data = gd.dfset1)
            plt.plot('xvalues', 'yvalues', color="#33a3cc", linewidth=3, label=cur_exer + " [Set 2]", data = gd.dfset2)
            plt.plot('xvalues', 'yvalues', color="#09046c", linewidth=3, label=cur_exer + " [Set 3]", data = gd.dfset3)
            plt.xlabel("Nr.")
            plt.ylabel("Reps")
            plt.legend(loc='upper left')
            plt.grid(True)
            plt.tight_layout()
        elif cur_exer == "Reverse Flies":
            gd.exer4()
            graph_exer = plt.figure(figsize=(15,4))
            graph_exer.patch.set_facecolor(background)
            line = FigureCanvasTkAgg(graph_exer, workout_1_graph)
            line.get_tk_widget().grid(columnspan=2, rowspan=1, column=0, row=1, padx=10, pady=10)
            plt.plot('xvalues', 'yvalues', color="#8380c2", linewidth=3, label=cur_exer + " [Set 1]", data = gd.dfset1)
            plt.plot('xvalues', 'yvalues', color="#33a3cc", linewidth=3, label=cur_exer + " [Set 2]", data = gd.dfset2)
            plt.plot('xvalues', 'yvalues', color="#09046c", linewidth=3, label=cur_exer + " [Set 3]", data = gd.dfset3)
            plt.xlabel("Nr.")
            plt.ylabel("Reps")
            plt.legend(loc='upper left')
            plt.grid(True)
            plt.tight_layout()
        elif cur_exer == "Push-Up":
            gd.exer5()
            graph_exer = plt.figure(figsize=(15,4))
            graph_exer.patch.set_facecolor(background)
            line = FigureCanvasTkAgg(graph_exer, workout_1_graph)
            line.get_tk_widget().grid(columnspan=2, rowspan=1, column=0, row=1, padx=10, pady=10)
            plt.plot('xvalues', 'yvalues', color="#8380c2", linewidth=3, label=cur_exer + " [Set 1]", data = gd.dfset1)
            plt.plot('xvalues', 'yvalues', color="#33a3cc", linewidth=3, label=cur_exer + " [Set 2]", data = gd.dfset2)
            plt.plot('xvalues', 'yvalues', color="#09046c", linewidth=3, label=cur_exer + " [Set 3]", data = gd.dfset3)
            plt.xlabel("Nr.")
            plt.ylabel("Reps")
            plt.legend(loc='upper left')
            plt.grid(True)
            plt.tight_layout()
        elif cur_exer == "Bicep Curls":
            gd.exer6()
            graph_exer = plt.figure(figsize=(15,4))
            graph_exer.patch.set_facecolor(background)
            line = FigureCanvasTkAgg(graph_exer, workout_1_graph)
            line.get_tk_widget().grid(columnspan=2, rowspan=1, column=0, row=1, padx=10, pady=10)
            plt.plot('xvalues', 'yvalues', color="#8380c2", linewidth=3, label=cur_exer + " [Set 1]", data = gd.dfset1)
            plt.plot('xvalues', 'yvalues', color="#33a3cc", linewidth=3, label=cur_exer + " [Set 2]", data = gd.dfset2)
            plt.plot('xvalues', 'yvalues', color="#09046c", linewidth=3, label=cur_exer + " [Set 3]", data = gd.dfset3)
            plt.xlabel("Nr.")
            plt.ylabel("Reps")
            plt.legend(loc='upper left')
            plt.grid(True)
            plt.tight_layout()
        elif cur_exer == "Trizep Extension":
            gd.exer7()
            graph_exer = plt.figure(figsize=(15,4))
            graph_exer.patch.set_facecolor(background)
            line = FigureCanvasTkAgg(graph_exer, workout_1_graph)
            line.get_tk_widget().grid(columnspan=2, rowspan=1, column=0, row=1, padx=10, pady=10)
            plt.plot('xvalues', 'yvalues', color="#8380c2", linewidth=3, label=cur_exer + " [Set 1]", data = gd.dfset1)
            plt.plot('xvalues', 'yvalues', color="#33a3cc", linewidth=3, label=cur_exer + " [Set 2]", data = gd.dfset2)
            plt.plot('xvalues', 'yvalues', color="#09046c", linewidth=3, label=cur_exer + " [Set 3]", data = gd.dfset3)
            plt.xlabel("Nr.")
            plt.ylabel("Reps")
            plt.legend(loc='upper left')
            plt.grid(True)
            plt.tight_layout()

    # Option für Übungen
    variable = tk.StringVar(workout_1_graph)
    variable.set(gd.exercise[5])
    variable.trace("w", change_Graph())
    opt = tk.OptionMenu(workout_1_graph, variable, *gd.exercise)
    opt.grid(columnspan=1, rowspan=1, column=1, row=0, padx=20, pady=20, sticky="W")
    opt.config(bg=btncolor, fg="white", font=("Calibri bold", 12), width=30, relief="groove", )

    # Select
    select = tk.Label(workout_1_graph, text="Select: ", relief="groove")
    select.grid(columnspan=1, rowspan=1, column=0, row=0, padx=20, pady=20, sticky="E")
    select.config(bg=btncolor, fg="white", font=("Calibri bold", 12), width=30, relief="groove")


# Alles für das 2. Workout
def workout_2():
    workout_2 = Toplevel(root)
    workout_2.title("Ring Workout")
    canvas_1 = tk.Canvas(workout_2, width=2440, height=1200, bg=background)
    canvas_1.grid(columnspan=2, rowspan=2)
    workout_2.geometry("+50+50")

root.mainloop()