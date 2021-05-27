import openpyxl
from tkinter import *
from tkinter.ttk import *
import os
import pandas as pd

workbook = openpyxl.load_workbook("Workout.xlsx")

plan = {
    1: "Übung_1",
    2: "Übung_2",
    3: "Übung_3",
    4: "Übung_4",
    5: "Übung_5",
    6: "Übung_6",
    7: "Übung_7"
}
# Aktuelle Übung
def show_exer():
    global exercise
    exercise = []
    # open sheet 1-7 für Übungsnamen
    for i in range(1, 8):
        sheet = workbook[plan[i]]
        if sheet.cell(row=1, column=8).value != None:
            exercise.append(sheet.cell(row=1, column=8).value)

show_exer()
print(exercise)