import openpyxl
from tkinter import *
from tkinter.ttk import *
import os
import pandas as pd


os.chdir(r"C:\Python\Projects\FitnessTracker")

workbook = openpyxl.load_workbook("Workout.xlsx")
sheet1 = workbook["Übung_1"]
sheet2 = workbook["Übung_2"]
sheet3 = workbook["Übung_3"]
sheet4 = workbook["Übung_4"]        
sheet5 = workbook["Übung_5"]
sheet6 = workbook["Übung_6"]
sheet7 = workbook["Übung_7"]
global plan
plan = {
    1: "Übung_1",
    2: "Übung_2",
    3: "Übung_3",
    4: "Übung_4",
    5: "Übung_5",
    6: "Übung_6",
    7: "Übung_7"
}

def exer1():
    global dfset1
    global dfset2
    global dfset3

    # Data import from Excel easy
    nrvalues = []
    start = 2
    for i in range(start, 1000):
        if sheet1.cell(row=i, column=1).value != None:
            nrvalues.append(sheet1.cell(row=i, column=1).value)
            start += 1

    set1 = []
    start = 2
    for i in range(start, 1000):
        if sheet1.cell(row=i, column=3).value != None:
            set1.append(sheet1.cell(row=i, column=3).value)
            start += 1

    set2 = []
    start = 2
    for i in range(start, 1000):
        if sheet1.cell(row=i, column=4).value != None:
            set2.append(sheet1.cell(row=i, column=4).value)
            start += 1 

    set3 = []
    start = 2
    for i in range(start, 1000):
        if sheet1.cell(row=i, column=5).value != None:
            set3.append(sheet1.cell(row=i, column=5).value)
            start += 1 
            

    # Data
    dfset1 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set1})
    dfset2 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set2})
    dfset3 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set3})


def exer2():
    global dfset1
    global dfset2
    global dfset3

    # Data import from Excel easy
    nrvalues = []
    start = 2
    for i in range(start, 1000):
        if sheet2.cell(row=i, column=1).value != None:
            nrvalues.append(sheet2.cell(row=i, column=1).value)
            start += 1

    set1 = []
    start = 2
    for i in range(start, 1000):
        if sheet2.cell(row=i, column=3).value != None:
            set1.append(sheet2.cell(row=i, column=3).value)
            start += 1

    set2 = []
    start = 2
    for i in range(start, 1000):
        if sheet2.cell(row=i, column=4).value != None:
            set2.append(sheet2.cell(row=i, column=4).value)
            start += 1 

    set3 = []
    start = 2
    for i in range(start, 1000):
        if sheet2.cell(row=i, column=5).value != None:
            set3.append(sheet2.cell(row=i, column=5).value)
            start += 1 
            

    # Data
    dfset1 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set1})
    dfset2 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set2})
    dfset3 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set3})

def exer3():
    global dfset1
    global dfset2
    global dfset3
    
    # Data import from Excel easy
    nrvalues = []
    start = 2
    for i in range(start, 1000):
        if sheet3.cell(row=i, column=1).value != None:
            nrvalues.append(sheet3.cell(row=i, column=1).value)
            start += 1

    set1 = []
    start = 2
    for i in range(start, 1000):
        if sheet3.cell(row=i, column=3).value != None:
            set1.append(sheet3.cell(row=i, column=3).value)
            start += 1

    set2 = []
    start = 2
    for i in range(start, 1000):
        if sheet3.cell(row=i, column=4).value != None:
            set2.append(sheet3.cell(row=i, column=4).value)
            start += 1 

    set3 = []
    start = 2
    for i in range(start, 1000):
        if sheet3.cell(row=i, column=5).value != None:
            set3.append(sheet3.cell(row=i, column=5).value)
            start += 1 

def exer4():
    global dfset1
    global dfset2
    global dfset3
    
    # Data import from Excel easy
    nrvalues = []
    start = 2
    for i in range(start, 1000):
        if sheet4.cell(row=i, column=1).value != None:
            nrvalues.append(sheet4.cell(row=i, column=1).value)
            start += 1

    set1 = []
    start = 2
    for i in range(start, 1000):
        if sheet4.cell(row=i, column=3).value != None:
            set1.append(sheet4.cell(row=i, column=3).value)
            start += 1

    set2 = []
    start = 2
    for i in range(start, 1000):
        if sheet4.cell(row=i, column=4).value != None:
            set2.append(sheet4.cell(row=i, column=4).value)
            start += 1 

    set3 = []
    start = 2
    for i in range(start, 1000):
        if sheet4.cell(row=i, column=5).value != None:
            set3.append(sheet4.cell(row=i, column=5).value)
            start += 1 
            

    # Data
    dfset1 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set1})
    dfset2 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set2})
    dfset3 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set3})

    # Data
    dfset1 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set1})
    dfset2 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set2})
    dfset3 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set3})

def exer5():
    global dfset1
    global dfset2
    global dfset3
    
    # Data import from Excel easy
    nrvalues = []
    start = 2
    for i in range(start, 1000):
        if sheet5.cell(row=i, column=1).value != None:
            nrvalues.append(sheet5.cell(row=i, column=1).value)
            start += 1

    set1 = []
    start = 2
    for i in range(start, 1000):
        if sheet5.cell(row=i, column=3).value != None:
            set1.append(sheet5.cell(row=i, column=3).value)
            start += 1

    set2 = []
    start = 2
    for i in range(start, 1000):
        if sheet5.cell(row=i, column=4).value != None:
            set2.append(sheet5.cell(row=i, column=4).value)
            start += 1 

    set3 = []
    start = 2
    for i in range(start, 1000):
        if sheet5.cell(row=i, column=5).value != None:
            set3.append(sheet5.cell(row=i, column=5).value)
            start += 1 
            

    # Data
    dfset1 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set1})
    dfset2 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set2})
    dfset3 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set3})

def exer6():
    global dfset1
    global dfset2
    global dfset3
    
    # Data import from Excel easy
    nrvalues = []
    start = 2
    for i in range(start, 1000):
        if sheet6.cell(row=i, column=1).value != None:
            nrvalues.append(sheet6.cell(row=i, column=1).value)
            start += 1

    set1 = []
    start = 2
    for i in range(start, 1000):
        if sheet6.cell(row=i, column=3).value != None:
            set1.append(sheet6.cell(row=i, column=3).value)
            start += 1

    set2 = []
    start = 2
    for i in range(start, 1000):
        if sheet6.cell(row=i, column=4).value != None:
            set2.append(sheet6.cell(row=i, column=4).value)
            start += 1 

    set3 = []
    start = 2
    for i in range(start, 1000):
        if sheet6.cell(row=i, column=5).value != None:
            set3.append(sheet6.cell(row=i, column=5).value)
            start += 1 

    dfset1 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set1})
    dfset2 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set2})
    dfset3 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set3})
            
def exer7():
    global dfset1
    global dfset2
    global dfset3
    
    # Data import from Excel easy
    nrvalues = []
    start = 2
    for i in range(start, 1000):
        if sheet7.cell(row=i, column=1).value != None:
            nrvalues.append(sheet7.cell(row=i, column=1).value)
            start += 1

    set1 = []
    start = 2
    for i in range(start, 1000):
        if sheet7.cell(row=i, column=3).value != None:
            set1.append(sheet7.cell(row=i, column=3).value)
            start += 1

    set2 = []
    start = 2
    for i in range(start, 1000):
        if sheet7.cell(row=i, column=4).value != None:
            set2.append(sheet7.cell(row=i, column=4).value)
            start += 1 

    set3 = []
    start = 2
    for i in range(start, 1000):
        if sheet7.cell(row=i, column=5).value != None:
            set3.append(sheet7.cell(row=i, column=5).value)
            start += 1 
            

    # Data
    dfset1 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set1})
    dfset2 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set2})
    dfset3 = pd.DataFrame({'xvalues': nrvalues, 'yvalues': set3})



# Aktuelle Übung
def show_exer():
    global exercise
    exercise = []
    # open sheet 1-7 für Übungsnamen
    for i in range(1, 8):
        sheet = workbook[plan[i]]
        if sheet.cell(row=1, column=8).value != None:
            exercise.append(sheet.cell(row=1, column=8).value)

show_exer()
